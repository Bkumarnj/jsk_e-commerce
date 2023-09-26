import calendar
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from itertools import cycle

# Define shifts and shift durations
shifts = [
    {"name": "Swing", "start_time": "05:30", "end_time": "15:00"},
    {"name": "Grave", "start_time": "13:30", "end_time": "23:00"},
    {"name": "Day", "start_time": "21:30", "end_time": "07:00"},
]

# Define the team of 6 people
team = ["Person 1", "Person 2", "Person 3", "Person 4", "Person 5", "Person 6"]

# Create a DataFrame to store the roster
roster = pd.DataFrame(columns=["Date", "Shift", "Primary", "Secondary"])

# Create a cycle of shifts to repeat through the year
shift_cycle = cycle(shifts)

# Define a function to assign shifts to team members
def assign_shifts(team, start_date, end_date):
    for date, shift in zip(pd.date_range(start_date, end_date), shift_cycle):
        for i in range(2 if date.weekday() < 5 else 1):
            primary = team.pop(0)
            secondary = team.pop(0)
            team.append(primary)
            team.append(secondary)
            roster.loc[len(roster)] = [date, shift["name"], primary, secondary]

# Generate the roster for each month of the year
for month in range(1, 13):
    _, last_day = calendar.monthrange(2023, month)
    start_date = f"2023-{month:02d}-01"
    end_date = f"2023-{month:02d}-{last_day:02d}"
    assign_shifts(team.copy(), start_date, end_date)

# Export the roster to an Excel file
wb = openpyxl.Workbook()
ws = wb.active

# Add header
header = ["Date", "Shift", "Primary", "Secondary"]
for col_num, header_text in enumerate(header, 1):
    cell = ws.cell(row=1, column=col_num)
    cell.value = header_text

# Append data to the Excel sheet
for r_idx, row in enumerate(dataframe_to_rows(roster, index=False), 2):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx)
        cell.value = value

# Save the Excel file
wb.save("shift-roster.xlsx")
